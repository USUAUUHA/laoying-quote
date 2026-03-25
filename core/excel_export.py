from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .calculator import calc_quote_row, migrate_dataframe
from .config import (
    EXPORT_FOLDER_SUFFIX,
    EXPORT_LY_NAME_CORE,
    EXPORT_REFERENCE_ROOT,
    NUMERIC_INPUT_COLS,
    SEAL_EXPORT_HEIGHT_PX,
    SEAL_EXPORT_WIDTH_PX,
    SEAL_IMAGE_ALT,
    SEAL_IMAGE_PATH,
    SURFACE_OPTIONS,
)

try:
    from openpyxl.drawing.image import Image as XLImage
except ImportError:
    XLImage = None  # type: ignore[assignment]


def _ensure_directory_writable(directory: Path) -> None:
    try:
        directory.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        raise RuntimeError(f"无法创建导出目录：{directory}") from e

    probe = directory / ".__write_test__.tmp"
    try:
        probe.write_text("ok", encoding="utf-8")
        probe.unlink(missing_ok=True)
    except Exception as e:
        raise RuntimeError(f"导出目录不可写：{directory}") from e


def _export_folder_today() -> Path:
    now = datetime.now()
    sub = f"{now.month}.{now.day}{EXPORT_FOLDER_SUFFIX}"
    folder = EXPORT_REFERENCE_ROOT / sub
    _ensure_directory_writable(folder)
    return folder


def _find_seal_image() -> Optional[Path]:
    for p in (SEAL_IMAGE_PATH, SEAL_IMAGE_ALT):
        if p.is_file():
            return p
    return None


_FILL_YELLOW = "FFF2CC"
_FILL_PINK = "FCE4D6"
_FILL_TEAL = "B7DEE8"
_FILL_HEADER = "D9E1F2"


def _surface_fee_header(surface: str) -> str:
    if surface == "无":
        return "表面处理费（无表面处理成品）"
    return f"{surface}表面处理费（成品）"


def compute_ly_export_breakdown(
    row: pd.Series,
    weight_coef: float,
    price_params: Dict[str, float],
) -> Optional[Dict[str, Any]]:
    calc_res, errs = calc_quote_row(row, weight_coef=weight_coef, price_params=price_params)
    if errs:
        return None

    nums: Dict[str, float] = {}
    for col in NUMERIC_INPUT_COLS:
        try:
            nums[col] = float(row.get(col, 0.0) or 0.0)
        except (TypeError, ValueError):
            nums[col] = 0.0

    return {
        "料号": str(row.get("料号", "") or "").strip(),
        "料品名称": str(row.get("料品名称", "") or "").strip(),
        "料品规格": str(row.get("料品规格", "") or "").strip(),
        "材质": str(calc_res["_material_name"]),
        "料费": round(float(calc_res["_material_cost"]), 2),
        "激光切割": round(float(calc_res["_cutting_cost"]), 2),
        "打磨": round(float(calc_res["打磨(元)"]), 2),
        "穿孔单价": round(float(calc_res["_pierce_unit"]), 4),
        "穿孔数量": nums["穿孔数量"],
        "激光+穿孔": round(float(calc_res["_cutting_cost"]) + float(calc_res["_pierce_cost"]), 2),
        "焊接总价": round(float(calc_res["_welding_cost"]), 2),
        "沉孔单价": round(float(calc_res["_countersink_unit"]), 4),
        "沉孔数量": nums["沉孔数量"],
        "沉孔费": round(float(calc_res["_countersink_cost"]), 2),
        "折弯单价": round(float(calc_res["_bend_unit_price"]), 4),
        "折弯数量": float(calc_res["_bend_qty_total"]),
        "折弯费": round(float(calc_res["_bend_cost"]), 2),
        "压铆单价": round(float(price_params["pem_a"]), 4),
        "压铆标准件费": round(float(price_params["pem_b"]) * nums["压铆数量"], 2),
        "压铆数量": nums["压铆数量"],
        "压铆费": round(float(calc_res["_pem_cost"]), 2),
        "其他标准件数量": 0,
        "其他标准件费": 0.0,
        "攻丝单价": round(float(calc_res["_tapping_unit_price"]), 4),
        "攻丝数量": float(calc_res["_tapping_qty_total"]),
        "攻丝费": round(float(calc_res["_tapping_cost"]), 2),
        "激光打标费": round(nums["激光打标"], 2),
        "_surface": str(calc_res["_surface"]),
        "_surface_fee": round(float(calc_res["_surface_cost"]), 2),
        "订单数量": int(calc_res["_qty"]),
        "合计": round(float(calc_res["含税总价(元)"]), 2),
    }


def _unique_surfaces_in_df(df: pd.DataFrame) -> List[str]:
    seen = set()
    order: List[str] = []
    for _, row in df.iterrows():
        seq = str(row.get("序号", ""))
        if seq.startswith("🔴"):
            continue
        su = str(row.get("表面处理", "")).strip()
        if su and su in SURFACE_OPTIONS and su not in seen:
            seen.add(su)
            order.append(su)
    return order


def export_table_excel(
    df: pd.DataFrame,
    weight_coef: float,
    price_params: Dict[str, float],
) -> Path:
    folder = _export_folder_today()
    now = datetime.now()
    fname = f"{now.month}.{now.day}-{EXPORT_LY_NAME_CORE}_{now.strftime('%H%M%S')}.xlsx"
    out_path = folder / fname

    work = migrate_dataframe(df)
    work["序号"] = work["序号"].astype(str).str.replace("🔴", "", regex=False)

    surfaces = _unique_surfaces_in_df(work) or ["无"]
    fixed_headers = [
        "序号",
        "料号",
        "料品名称",
        "料品规格",
        "材质",
        "料费",
        "激光切割",
        "打磨",
        "穿孔单价",
        "穿孔数量",
        "激光+穿孔",
        "焊接总价",
        "沉孔单价",
        "沉孔数量",
        "沉孔费",
        "折弯单价",
        "折弯数量",
        "折弯费",
        "压铆单价",
        "压铆标准件费",
        "压铆数量",
        "压铆费",
        "其他标准件数量",
        "其他标准件费",
        "攻丝单价",
        "攻丝数量",
        "攻丝费",
        "激光打标费",
    ]
    surface_headers = [_surface_fee_header(s) for s in surfaces]
    headers = fixed_headers + surface_headers + ["订单数量", "合计"]

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    yellow_titles = {
        "料费",
        "激光切割",
        "打磨",
        "穿孔单价",
        "折弯单价",
        "压铆单价",
        "攻丝单价",
        "激光打标费",
        "订单数量",
    }
    fill_y = PatternFill(fill_type="solid", fgColor=_FILL_YELLOW)
    fill_p = PatternFill(fill_type="solid", fgColor=_FILL_PINK)
    fill_t = PatternFill(fill_type="solid", fgColor=_FILL_TEAL)
    fill_h = PatternFill(fill_type="solid", fgColor=_FILL_HEADER)

    wb = Workbook()
    ws = wb.active
    ws.title = "报价明细"

    ncols = len(headers)
    title_text = f"{now.month}.{now.day}-{EXPORT_LY_NAME_CORE}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    tcell = ws.cell(row=1, column=1, value=title_text)
    tcell.font = Font(bold=True, size=14)
    tcell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        if h in yellow_titles:
            cell.fill = fill_y
        elif h == "合计":
            cell.fill = fill_t
        elif h in surface_headers or "表面处理费" in str(h) or "成品" in str(h):
            cell.fill = fill_p
        else:
            cell.fill = fill_h

    r_data = 3
    sum_total = 0.0
    for _, row in work.iterrows():
        seq = str(row.get("序号", ""))
        if seq.startswith("🔴"):
            continue

        breakdown = compute_ly_export_breakdown(row, weight_coef, price_params)
        if breakdown is None:
            values = [seq] + [""] * (len(headers) - 1)
        else:
            values = [seq]
            values.extend(breakdown[k] for k in fixed_headers[1:])
            for surface in surfaces:
                values.append(breakdown["_surface_fee"] if breakdown["_surface"] == surface else 0)
            values.append(breakdown["订单数量"])
            values.append(breakdown["合计"])
            sum_total += float(breakdown["合计"])

        for c_idx, value in enumerate(values, 1):
            cell = ws.cell(row=r_data, column=c_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
        r_data += 1

    total_row = r_data
    ws.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
    ws.cell(row=total_row, column=1).border = border
    total_col = headers.index("合计") + 1
    sum_cell = ws.cell(row=total_row, column=total_col, value=round(sum_total, 2))
    sum_cell.font = Font(bold=True)
    sum_cell.fill = fill_t
    sum_cell.border = border
    for c_idx in range(2, total_col):
        ws.cell(row=total_row, column=c_idx, value="").border = border

    seal_anchor_col = ncols + 2
    img_path = _find_seal_image()
    if img_path and XLImage is not None:
        try:
            img = XLImage(str(img_path))
            img.width = SEAL_EXPORT_WIDTH_PX
            img.height = SEAL_EXPORT_HEIGHT_PX
            ws.add_image(img, f"{get_column_letter(seal_anchor_col)}{total_row}")
        except Exception:
            pass

    for extra in (ncols + 1, ncols + 2, ncols + 3):
        ws.column_dimensions[get_column_letter(extra)].width = 14
    for idx in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 12

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 36

    try:
        wb.save(out_path)
    except Exception as e:
        raise RuntimeError(f"导出文件保存失败：{out_path}") from e
    return out_path
