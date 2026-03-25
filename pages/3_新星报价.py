"""
新星客户专属钣金报价模板

说明：
- 本文件仅实现新星客户模板，不改动其他客户页面；
- 保留批量粘贴、实时计算、Excel 导出能力；
- 新增焊接标准件、攻丝模块，并按要求加入长/宽自动 +10 规则。
"""

from __future__ import annotations

import sys
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# 项目根目录，便于加载公共导出模块
_ROOT = Path(__file__).resolve().parents[1]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))
from common_quote_export import build_xinxing_quote_detail_workbook_bytes

st.set_page_config(page_title="新星报价", page_icon="🧩", layout="wide")

# ==================== 全局配置区（可直接修改） ====================
# 新增材质参数（密度g/cm³，单价元/kg）
MATERIAL_PARAMS: Dict[str, Dict[str, float]] = {
    "Q235冷轧钢板": {"density": 7.85, "price": 5},
    "镀锌板": {"density": 7.85, "price": 5},
    "5052铝板": {"density": 2.73, "price": 24},
    "SUS304不锈钢": {"density": 7.93, "price": 18},
}
# 新增加工费固定单价
WELD_STD_PART_PRICE = 0.8  # 焊接标准件单价（元/个）
TAPPING_PRICE = {"铝板": 0.4, "其他": 0.6}  # 攻丝单价（元/个）
WELD_PRICE = {"连续焊": 0.05, "点焊": 0.3}  # 焊接单价（元/mm/个）

LOSS_RATE = 0.05
TAX_RATE = 0.13
DEFAULT_ROWS = 8
SHEET_NAME = "Sheet1"
EXPORT_NAME_PREFIX = "新星-钣金报价"
WATERMARK_TEXT = "青岛宏泰铭润机械"

MATERIAL_OPTIONS = ["Q235冷轧钢板", "镀锌板", "5052铝板", "SUS304不锈钢"]
WELD_METHOD_OPTIONS = ["无焊接", "连续焊", "点焊"]
WELD_PART_OPTIONS = ["无", "M5螺母", "M6螺栓"]
TAP_SPEC_OPTIONS = ["无攻丝", "M3", "M4", "M5"]
SURFACE_OPTIONS = ["无", "喷粉", "磷化"]

INPUT_COLS = [
    "料号",
    "物料名称",
    "规格型号",
    "材质",
    "长(mm)",
    "宽(mm)",
    "厚(mm)",
    "数量",
    "切割米数(m)",
    "折弯次数",
    "焊接方式",
    "焊接长度(mm)/焊点数量",
    "焊接标准件名称",
    "焊接标准件数量",
    "攻丝规格",
    "攻丝数量",
    "表面处理类型",
    "备注",
]

CALC_COLS = [
    "切割费(元)",
    "折弯费(元)",
    "焊接费(元)",
    "焊接标准件费(元)",
    "攻丝费(元)",
    "表面处理费(元)",
    "单重(kg)",
    "总重(kg)",
    "材料费(元)",
    "单品含税总价(元)",
    "实际长(mm)",
    "实际宽(mm)",
    "焊接输入提示",
]


def _to_float(v: Any, default: float = 0.0) -> float:
    try:
        if v is None or str(v).strip() == "":
            return default
        return float(v)
    except Exception:
        return default


def _to_int(v: Any, default: int = 0) -> int:
    try:
        if v is None or str(v).strip() == "":
            return default
        return int(float(v))
    except Exception:
        return default


def _match_material(text: Any) -> str:
    t = str(text or "").strip()
    if not t:
        return "Q235冷轧钢板"
    if t in MATERIAL_PARAMS:
        return t
    lower = t.lower()
    if "镀锌" in t:
        return "镀锌板"
    if "5052" in lower or "铝" in t:
        return "5052铝板"
    if "304" in lower or "sus304" in lower or "不锈钢" in t:
        return "SUS304不锈钢"
    if "q235" in lower or "冷轧" in t:
        return "Q235冷轧钢板"
    return "Q235冷轧钢板"


def _init_df(n: int = DEFAULT_ROWS) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    for i in range(n):
        rows.append(
            {
                "序号": i + 1,
                "料号": "",
                "物料名称": "",
                "规格型号": "",
                "材质": "Q235冷轧钢板",
                "长(mm)": 0.0,
                "宽(mm)": 0.0,
                "厚(mm)": 1.0,
                "数量": 1,
                "切割米数(m)": 0.0,
                "切割费(元)": 0.0,
                "折弯次数": 0,
                "折弯费(元)": 0.0,
                "焊接方式": "无焊接",
                "焊接长度(mm)/焊点数量": 0.0,
                "焊接费(元)": 0.0,
                "焊接标准件名称": "无",
                "焊接标准件数量": 0,
                "焊接标准件费(元)": 0.0,
                "攻丝规格": "无攻丝",
                "攻丝数量": 0,
                "攻丝费(元)": 0.0,
                "表面处理类型": "无",
                "表面处理费(元)": 0.0,
                "单重(kg)": 0.0,
                "总重(kg)": 0.0,
                "材料费(元)": 0.0,
                "单品含税总价(元)": 0.0,
                "备注": "",
                "实际长(mm)": 0.0,
                "实际宽(mm)": 0.0,
                "焊接输入提示": "",
            }
        )
    return pd.DataFrame(rows)


def _parse_batch_text(text: str) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for line in [x for x in text.splitlines() if x.strip()]:
        parts = [p.strip() for p in line.split("\t")]
        row = {
            "料号": parts[0] if len(parts) > 0 else "",
            "物料名称": parts[1] if len(parts) > 1 else "",
            "规格型号": parts[2] if len(parts) > 2 else "",
            "材质": _match_material(parts[3] if len(parts) > 3 else ""),
            "长(mm)": _to_float(parts[4] if len(parts) > 4 else 0),
            "宽(mm)": _to_float(parts[5] if len(parts) > 5 else 0),
            "厚(mm)": _to_float(parts[6] if len(parts) > 6 else 1.0, 1.0),
            "数量": _to_int(parts[7] if len(parts) > 7 else 1, 1),
            "切割米数(m)": _to_float(parts[8] if len(parts) > 8 else 0),
            "折弯次数": _to_int(parts[9] if len(parts) > 9 else 0),
            "焊接方式": parts[10] if len(parts) > 10 and parts[10] in WELD_METHOD_OPTIONS else "无焊接",
            "焊接长度(mm)/焊点数量": _to_float(parts[11] if len(parts) > 11 else 0),
            "焊接标准件名称": parts[12] if len(parts) > 12 and parts[12] else "无",
            "焊接标准件数量": _to_int(parts[13] if len(parts) > 13 else 0),
            "攻丝规格": parts[14] if len(parts) > 14 and parts[14] else "无攻丝",
            "攻丝数量": _to_int(parts[15] if len(parts) > 15 else 0),
            "表面处理类型": parts[16] if len(parts) > 16 and parts[16] in SURFACE_OPTIONS else "无",
            "备注": parts[17] if len(parts) > 17 else "",
        }
        # 支持在“长(mm)”里直接粘贴 500x300 / 500×300
        if row["宽(mm)"] <= 0 and isinstance(parts[4] if len(parts) > 4 else "", str):
            size_raw = parts[4] if len(parts) > 4 else ""
            if "x" in size_raw.lower() or "×" in size_raw:
                token = size_raw.lower().replace("×", "x").split("x")
                if len(token) >= 2:
                    row["长(mm)"] = _to_float(token[0], 0)
                    row["宽(mm)"] = _to_float(token[1], 0)
        out.append(row)
    return out


def _calc_df(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["材质"] = out["材质"].apply(_match_material)
    out["数量"] = out["数量"].apply(lambda x: max(_to_int(x, 1), 1))
    out["厚(mm)"] = out["厚(mm)"].apply(lambda x: max(_to_float(x, 1.0), 0))
    out["长(mm)"] = out["长(mm)"].apply(lambda x: max(_to_float(x, 0), 0))
    out["宽(mm)"] = out["宽(mm)"].apply(lambda x: max(_to_float(x, 0), 0))
    out["切割米数(m)"] = out["切割米数(m)"].apply(lambda x: max(_to_float(x, 0), 0))
    out["折弯次数"] = out["折弯次数"].apply(lambda x: max(_to_int(x, 0), 0))
    out["焊接长度(mm)/焊点数量"] = out["焊接长度(mm)/焊点数量"].apply(lambda x: max(_to_float(x, 0), 0))
    out["焊接标准件数量"] = out["焊接标准件数量"].apply(lambda x: max(_to_int(x, 0), 0))
    out["攻丝数量"] = out["攻丝数量"].apply(lambda x: max(_to_int(x, 0), 0))

    for i, row in out.iterrows():
        material = _match_material(row.get("材质", ""))
        length_raw = _to_float(row.get("长(mm)", 0))
        width_raw = _to_float(row.get("宽(mm)", 0))
        thick = _to_float(row.get("厚(mm)", 1.0), 1.0)
        qty = max(_to_int(row.get("数量", 1), 1), 1)
        cut_m = _to_float(row.get("切割米数(m)", 0))
        bend_count = _to_int(row.get("折弯次数", 0))
        weld_method = row.get("焊接方式", "无焊接")
        weld_value = _to_float(row.get("焊接长度(mm)/焊点数量", 0))
        weld_std_qty = _to_int(row.get("焊接标准件数量", 0))
        tap_qty = _to_int(row.get("攻丝数量", 0))
        surface = row.get("表面处理类型", "无")

        actual_length = length_raw + 10 if length_raw > 0 else 0
        actual_width = width_raw + 10 if width_raw > 0 else 0

        if thick < 1:
            cut_fee = cut_m * 0.8
        else:
            cut_fee = cut_m * thick * 0.8

        bend_unit = 0.6 if (actual_length <= 100 and actual_width <= 100 and actual_length > 0 and actual_width > 0) else 0.7
        bend_fee = bend_count * bend_unit * qty

        if weld_method == "连续焊":
            weld_fee = weld_value * WELD_PRICE["连续焊"]
            weld_hint = "请输入焊接长度(mm)"
        elif weld_method == "点焊":
            weld_fee = weld_value * WELD_PRICE["点焊"]
            weld_hint = "请输入焊点数量"
        else:
            weld_fee = 0.0
            weld_hint = "无焊接，无需填写"

        weld_std_fee = weld_std_qty * WELD_STD_PART_PRICE * qty

        tap_unit = TAPPING_PRICE["铝板"] if "5052" in material or "铝" in material else TAPPING_PRICE["其他"]
        tap_fee = tap_qty * tap_unit * qty

        area = actual_length * actual_width * 2
        if surface == "喷粉":
            surface_fee = area / 1_000_000 * 25
        elif surface == "磷化":
            surface_fee = area / 1_000_000 * 15
        else:
            surface_fee = 0.0

        density = MATERIAL_PARAMS.get(material, MATERIAL_PARAMS["Q235冷轧钢板"])["density"]
        single_weight = actual_length * actual_width * thick * density / 1_000_000
        total_weight = single_weight * qty
        material_price = MATERIAL_PARAMS.get(material, MATERIAL_PARAMS["Q235冷轧钢板"])["price"]
        material_fee = total_weight * material_price * (1 + LOSS_RATE)

        total_price_tax = (material_fee + cut_fee + bend_fee + weld_fee + weld_std_fee + tap_fee + surface_fee) * (1 + TAX_RATE)

        out.at[i, "序号"] = i + 1
        out.at[i, "切割费(元)"] = round(cut_fee, 2)
        out.at[i, "折弯费(元)"] = round(bend_fee, 2)
        out.at[i, "焊接费(元)"] = round(weld_fee, 2)
        out.at[i, "焊接标准件费(元)"] = round(weld_std_fee, 2)
        out.at[i, "攻丝费(元)"] = round(tap_fee, 2)
        out.at[i, "表面处理费(元)"] = round(surface_fee, 2)
        out.at[i, "单重(kg)"] = round(single_weight, 4)
        out.at[i, "总重(kg)"] = round(total_weight, 4)
        out.at[i, "材料费(元)"] = round(material_fee, 2)
        out.at[i, "单品含税总价(元)"] = round(total_price_tax, 2)
        out.at[i, "实际长(mm)"] = round(actual_length, 2)
        out.at[i, "实际宽(mm)"] = round(actual_width, 2)
        out.at[i, "焊接输入提示"] = weld_hint
    return out


def _to_excel_bytes(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    title = f"{datetime.now():%Y-%m-%d} 新星客户钣金报价明细"
    headers = list(df.columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1, value=title).font = Font(size=14, bold=True, color="FFFFFF")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=1, column=1).fill = PatternFill(fill_type="solid", fgColor="1E3A8A")

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="1E3A8A")

    for r_idx, row in enumerate(df.itertuples(index=False), start=3):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    sum_row = len(df) + 3
    ws.cell(row=sum_row, column=1, value="汇总")
    price_col = headers.index("单品含税总价(元)") + 1
    ws.cell(row=sum_row, column=price_col, value=round(float(df["单品含税总价(元)"].sum()), 2))
    ws.cell(row=sum_row, column=1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=sum_row, column=1).fill = PatternFill(fill_type="solid", fgColor="1E3A8A")
    ws.cell(row=sum_row, column=price_col).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=sum_row, column=price_col).fill = PatternFill(fill_type="solid", fgColor="1E3A8A")

    # 列宽控制：标准件名称/攻丝规格要求宽度 12，其他按内容自动估算
    for idx, h in enumerate(headers, start=1):
        if h in ("焊接标准件名称", "攻丝规格"):
            width = 12
        else:
            sample = [str(h)] + [str(x) for x in df[h].head(50).tolist()]
            width = min(max(len(max(sample, key=len)) + 2, 10), 24)
        ws.column_dimensions[get_column_letter(idx)].width = width

    # 水印配置（保留）：放置在远列，避免影响正常阅读
    ws["AZ1"] = WATERMARK_TEXT
    ws["AZ1"].font = Font(color="D1D5DB", italic=True, size=10)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


st.title("新星客户专属钣金报价模板")

with st.expander("参数说明", expanded=False):
    st.write(
        "材质默认优先Q235冷轧钢板/镀锌板/SUS304不锈钢/5052铝板，支持关键词检索；"
        "长(mm)/宽(mm)输入后系统自动+10计算实际值；尺寸单位mm，重量单位kg，"
        "所有公式基于新星泡菜冰箱报价规则"
    )

if "xx_quote_df" not in st.session_state:
    st.session_state.xx_quote_df = _init_df()

with st.expander("批量粘贴（Tab 分隔，按列顺序）", expanded=False):
    text = st.text_area(
        "可粘贴：料号、物料名称、规格型号、材质、长、宽、厚、数量、切割米数、折弯次数、焊接方式、焊接长度/焊点、焊接标准件名称、焊接标准件数量、攻丝规格、攻丝数量、表面处理类型、备注",
        height=120,
    )
    c1, c2 = st.columns([1, 1])
    with c1:
        if st.button("覆盖导入粘贴数据", use_container_width=True):
            parsed = _parse_batch_text(text)
            if not parsed:
                st.warning("未识别到有效粘贴数据。")
            else:
                st.session_state.xx_quote_df = _calc_df(pd.DataFrame(parsed))
                st.success(f"已导入 {len(parsed)} 行。")
    with c2:
        if st.button("在末尾追加粘贴数据", use_container_width=True):
            parsed = _parse_batch_text(text)
            if not parsed:
                st.warning("未识别到有效粘贴数据。")
            else:
                merged = pd.concat([st.session_state.xx_quote_df, pd.DataFrame(parsed)], ignore_index=True)
                st.session_state.xx_quote_df = _calc_df(merged)
                st.success(f"已追加 {len(parsed)} 行。")

col_add, col_del = st.columns([1, 1])
with col_add:
    if st.button("新增一行", use_container_width=True):
        st.session_state.xx_quote_df = pd.concat([st.session_state.xx_quote_df, _init_df(1)], ignore_index=True)
with col_del:
    if st.button("删除最后一行", use_container_width=True) and len(st.session_state.xx_quote_df) > 1:
        st.session_state.xx_quote_df = st.session_state.xx_quote_df.iloc[:-1].reset_index(drop=True)

edited = st.data_editor(
    st.session_state.xx_quote_df,
    use_container_width=True,
    num_rows="dynamic",
    key="xx_editor",
    column_config={
        "序号": st.column_config.NumberColumn(disabled=True),
        "材质": st.column_config.SelectboxColumn(options=MATERIAL_OPTIONS, required=False),
        "焊接方式": st.column_config.SelectboxColumn(options=WELD_METHOD_OPTIONS, required=False),
        "焊接标准件名称": st.column_config.SelectboxColumn(options=WELD_PART_OPTIONS, required=False),
        "攻丝规格": st.column_config.SelectboxColumn(options=TAP_SPEC_OPTIONS, required=False),
        "表面处理类型": st.column_config.SelectboxColumn(options=SURFACE_OPTIONS, required=False),
        "切割费(元)": st.column_config.NumberColumn(disabled=True),
        "折弯费(元)": st.column_config.NumberColumn(disabled=True),
        "焊接费(元)": st.column_config.NumberColumn(disabled=True),
        "焊接标准件费(元)": st.column_config.NumberColumn(disabled=True),
        "攻丝费(元)": st.column_config.NumberColumn(disabled=True),
        "表面处理费(元)": st.column_config.NumberColumn(disabled=True),
        "单重(kg)": st.column_config.NumberColumn(disabled=True),
        "总重(kg)": st.column_config.NumberColumn(disabled=True),
        "材料费(元)": st.column_config.NumberColumn(disabled=True),
        "单品含税总价(元)": st.column_config.NumberColumn(disabled=True),
        "实际长(mm)": st.column_config.NumberColumn(disabled=True),
        "实际宽(mm)": st.column_config.NumberColumn(disabled=True),
        "焊接输入提示": st.column_config.TextColumn(disabled=True),
    },
)

st.session_state.xx_quote_df = _calc_df(pd.DataFrame(edited))
df_show = st.session_state.xx_quote_df

# 必填校验：仅保持材质、长、宽、厚、数量
invalid_mask = (
    df_show["材质"].astype(str).str.strip().eq("")
    | (df_show["长(mm)"].astype(float) <= 0)
    | (df_show["宽(mm)"].astype(float) <= 0)
    | (df_show["厚(mm)"].astype(float) <= 0)
    | (df_show["数量"].astype(float) <= 0)
)
invalid_count = int(invalid_mask.sum())
if invalid_count > 0:
    st.warning(f"有 {invalid_count} 行未满足必填条件（材质、长、宽、厚、数量），请检查。")

total_items = int(len(df_show))
total_weight = float(df_show["总重(kg)"].sum())
total_price = float(df_show["单品含税总价(元)"].sum())

s1, s2, s3 = st.columns(3)
s1.metric("零件行数", total_items)
s2.metric("整单总重(kg)", f"{total_weight:.3f}")
s3.metric("整单含税总价(元)", f"{total_price:.2f}")

st.caption("长/宽自动 +10 后参与全部计算，可直接查看“实际长(mm) / 实际宽(mm)”列。")

export_name = f"{EXPORT_NAME_PREFIX}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
excel_data = _to_excel_bytes(df_show)
st.download_button(
    "导出报价明细（Excel·含公式+单价参数表）",
    data=build_xinxing_quote_detail_workbook_bytes(
        df_show,
        material_params=MATERIAL_PARAMS,
        loss_rate=LOSS_RATE,
        tax_rate=TAX_RATE,
        weld_std_price=WELD_STD_PART_PRICE,
        tap_al=TAPPING_PRICE["铝板"],
        tap_other=TAPPING_PRICE["其他"],
        weld_cont=WELD_PRICE["连续焊"],
        weld_spot=WELD_PRICE["点焊"],
    ),
    file_name=export_name.replace(".xlsx", "_含公式.xlsx"),
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
    help="主表计算列为 Excel 公式，「单价参数」工作表可改税率/损耗/加工单价；另附材质对照与使用说明。",
)
st.download_button(
    "导出Excel报价文件（纯数值快照）",
    data=excel_data,
    file_name=export_name,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
    help="不含公式，适合快速归档或与旧流程对齐。",
)

st.divider()
st.page_link("报价系统首页.py", label="返回首页", icon="🏠")
