# -*- coding: utf-8 -*-
"""
各客户报价页共用的「报价明细」Excel 导出工具。

设计目标：
- 新星：主表写入可编辑的 Excel 公式，并附带「单价参数」工作表，改参数即可联动重算。
- 崂应：主表为当前界面数值快照（与网页「更新计算」结果一致），另附加工单价、材料库、公式文字说明，便于线下对照与改价。
- 占位页：导出空白模板与说明，便于后续接表。

说明：崂应脚本内部计算非常复杂，本模块不在 Excel 中 1:1 复刻全部公式，避免维护双份逻辑；
      明细数值与侧边栏单价参数已完整导出，满足「可改表格、可对账」需求。
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Mapping, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ----- 新星：单价参数表布局（B 列为数值，供公式引用） -----
_XX_PARAM_SHEET = "单价参数"
_XX_PARAM_ROWS: List[Tuple[str, float, str]] = [
    ("税率", 0.13, "单品含税总价 = (成本合计)×(1+税率)"),
    ("损耗率", 0.05, "材料费 = 总重×材料单价×(1+损耗率)"),
    ("切割系数", 0.8, "厚<1：切割米数×系数；否则 米数×厚×系数"),
    ("折弯单价_小件", 0.6, "实际长、宽均≤100 且>0 时，每次×该单价×数量"),
    ("折弯单价_大件", 0.7, "不满足小件条件时，每次×该单价×数量"),
    ("连续焊_每mm", 0.05, "连续焊：焊接长度×该单价"),
    ("点焊_每点", 0.3, "点焊：焊点数量×该单价"),
    ("焊接标准件_单价", 0.8, "标准件费 = 数量×单价×总数量"),
    ("攻丝单价_铝板", 0.4, "材质含 5052 或「铝」时按此单价"),
    ("攻丝单价_其他", 0.6, "其余材质攻丝单价"),
    ("喷粉_每平方米", 25, "表面处理：实际长×实际宽×2÷1e6×该值"),
    ("磷化_每平方米", 15, "同上"),
]

# 参数值所在行号（从第 2 行开始写标签，数值在 B 列）
def _xx_param_b_row(idx: int) -> int:
    """idx 为 _XX_PARAM_ROWS 的 0-based 下标，返回 Excel 行号。"""
    return 2 + idx


def _xx_param_ref(row_idx: int) -> str:
    """生成对单价参数表 B 列的绝对引用。"""
    return f"'{_XX_PARAM_SHEET}'!$B${row_idx}"


def build_xinxing_quote_detail_workbook_bytes(
    df: pd.DataFrame,
    *,
    title: str = "新星客户钣金报价明细（含公式）",
    material_params: Optional[Mapping[str, Dict[str, float]]] = None,
    loss_rate: Optional[float] = None,
    tax_rate: Optional[float] = None,
    weld_std_price: Optional[float] = None,
    tap_al: Optional[float] = None,
    tap_other: Optional[float] = None,
    weld_cont: Optional[float] = None,
    weld_spot: Optional[float] = None,
) -> bytes:
    """
    新星：导出带公式的明细表 + 单价参数表 + 材质单价密度对照 + 使用说明。

    主表列顺序与传入 df 一致；计算列写入公式，输入列写入当前快照值（可在 Excel 中改）。
    """
    wb = Workbook()

    # ---------- Sheet：单价参数 ----------
    ws_p = wb.active
    ws_p.title = _XX_PARAM_SHEET
    ws_p["A1"] = "项目"
    ws_p["B1"] = "数值（可修改）"
    ws_p["C1"] = "说明"
    override_by_name: Dict[str, Optional[float]] = {
        "税率": tax_rate,
        "损耗率": loss_rate,
        "焊接标准件_单价": weld_std_price,
        "攻丝单价_铝板": tap_al,
        "攻丝单价_其他": tap_other,
        "连续焊_每mm": weld_cont,
        "点焊_每点": weld_spot,
    }
    for i, (name, val, desc) in enumerate(_XX_PARAM_ROWS):
        r = _xx_param_b_row(i)
        ws_p.cell(row=r, column=1, value=name)
        ov = override_by_name.get(name)
        ws_p.cell(row=r, column=2, value=float(ov) if ov is not None else float(val))
        ws_p.cell(row=r, column=3, value=desc)
    hdr = Font(bold=True, color="FFFFFF")
    fill = PatternFill(fill_type="solid", fgColor="1E3A8A")
    for c in range(1, 4):
        cell = ws_p.cell(row=1, column=c)
        cell.font = hdr
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_p.column_dimensions["A"].width = 22
    ws_p.column_dimensions["B"].width = 14
    ws_p.column_dimensions["C"].width = 56

    # ---------- Sheet：材质密度与材料单价（便于增删材质时改表） ----------
    ws_m = wb.create_sheet("材质密度与单价")
    ws_m.append(["材质", "密度(g/cm³)", "材料单价(元/kg)", "说明"])
    mp = dict(material_params) if material_params else {}
    default_rows = [
        ("Q235冷轧钢板", 7.85, 5),
        ("镀锌板", 7.85, 5),
        ("5052铝板", 2.73, 24),
        ("SUS304不锈钢", 7.93, 18),
    ]
    for name, d, p in default_rows:
        row = mp.get(name, {"density": d, "price": p})
        ws_m.append([name, float(row.get("density", d)), float(row.get("price", p)), "与网页 MATERIAL_PARAMS 一致；可在此增行后自行改公式引用"])
    for c in range(1, 5):
        cell = ws_m.cell(row=1, column=c)
        cell.font = hdr
        cell.fill = fill
    ws_m.column_dimensions["A"].width = 18
    ws_m.column_dimensions["D"].width = 40

    # ---------- Sheet：报价明细（公式） ----------
    ws = wb.create_sheet("报价明细", 0)
    headers: List[str] = list(df.columns)
    ncols = len(headers)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    tcell = ws.cell(row=1, column=1, value=f"{datetime.now():%Y-%m-%d} {title}")
    tcell.font = Font(size=14, bold=True, color="FFFFFF")
    tcell.fill = fill
    tcell.alignment = Alignment(horizontal="center", vertical="center")

    col_index = {h: i + 1 for i, h in enumerate(headers)}
    _need = [
        "长(mm)",
        "宽(mm)",
        "厚(mm)",
        "材质",
        "数量",
        "切割米数(m)",
        "折弯次数",
        "焊接方式",
        "焊接长度(mm)/焊点数量",
        "焊接标准件数量",
        "攻丝数量",
        "表面处理类型",
        "切割费(元)",
        "折弯费(元)",
        "焊接费(元)",
        "焊接标准件费(元)",
        "攻丝费(元)",
        "表面处理费(元)",
        "单重(kg)",
        "总重(kg)",
        "材料费(元)",
        "单品含税总价(元)",
        "实际长(mm)",
        "实际宽(mm)",
        "焊接输入提示",
    ]
    for h in _need:
        if h not in col_index:
            raise ValueError(f"导出失败：明细表缺少列「{h}」，请保持新星模板列名不变。")
    L = lambda name: get_column_letter(col_index[name])  # noqa: E731

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = hdr
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_start = 3
    n = len(df)
    pr = {name: _xx_param_b_row(i) for i, (name, _, _) in enumerate(_XX_PARAM_ROWS)}

    def cref(key: str) -> str:
        row_map = {
            "税率": pr["税率"],
            "损耗": pr["损耗率"],
            "切": pr["切割系数"],
            "弯小": pr["折弯单价_小件"],
            "弯大": pr["折弯单价_大件"],
            "连": pr["连续焊_每mm"],
            "点": pr["点焊_每点"],
            "标": pr["焊接标准件_单价"],
            "丝铝": pr["攻丝单价_铝板"],
            "丝他": pr["攻丝单价_其他"],
            "喷": pr["喷粉_每平方米"],
            "磷": pr["磷化_每平方米"],
        }
        return _xx_param_ref(row_map[key])

    for r_off in range(n):
        excel_row = data_start + r_off
        row = df.iloc[r_off]

        # 输入列：写入快照值
        for h in headers:
            if h in (
                "切割费(元)",
                "折弯费(元)",
                "焊接费(元)",
                "焊接标准件费(元)",
                "攻丝费(元)",
                "表面处理费(元)",
                "单重(kg)",
                "总重(kg)",
                "材料费(元)",
                "单品含税总价(元)",
                "实际长(mm)",
                "实际宽(mm)",
                "焊接输入提示",
            ):
                continue
            if h == "序号":
                ws.cell(row=excel_row, column=col_index[h], value=f"=ROW()-{data_start - 1}")
                continue
            val = row.get(h, "")
            ws.cell(row=excel_row, column=col_index[h], value=val)

        F, G, H, E, I, J, Lc, N, O, R, U, W = (
            L("长(mm)"),
            L("宽(mm)"),
            L("厚(mm)"),
            L("材质"),
            L("数量"),
            L("切割米数(m)"),
            L("折弯次数"),
            L("焊接方式"),
            L("焊接长度(mm)/焊点数量"),
            L("焊接标准件数量"),
            L("攻丝数量"),
            L("表面处理类型"),
        )
        AD, AE, K, M, P, S, V, X, Y, Z, AA, AB, AF = (
            L("实际长(mm)"),
            L("实际宽(mm)"),
            L("切割费(元)"),
            L("折弯费(元)"),
            L("焊接费(元)"),
            L("焊接标准件费(元)"),
            L("攻丝费(元)"),
            L("表面处理费(元)"),
            L("单重(kg)"),
            L("总重(kg)"),
            L("材料费(元)"),
            L("单品含税总价(元)"),
            L("焊接输入提示"),
        )

        ws.cell(row=excel_row, column=col_index["实际长(mm)"], value=f"=IF({F}{excel_row}>0,{F}{excel_row}+10,0)")
        ws.cell(row=excel_row, column=col_index["实际宽(mm)"], value=f"=IF({G}{excel_row}>0,{G}{excel_row}+10,0)")
        ws.cell(row=excel_row, column=col_index["切割费(元)"], value=f"=IF({H}{excel_row}<1,{J}{excel_row}*{cref('切')},{J}{excel_row}*{H}{excel_row}*{cref('切')})")
        ws.cell(row=excel_row, column=col_index["折弯费(元)"], value=(
            f"=IF(AND({AD}{excel_row}<=100,{AE}{excel_row}<=100,{AD}{excel_row}>0,{AE}{excel_row}>0),"
            f"{Lc}{excel_row}*{cref('弯小')}*{I}{excel_row},{Lc}{excel_row}*{cref('弯大')}*{I}{excel_row})"
        ))
        ws.cell(row=excel_row, column=col_index["焊接费(元)"], value=(
            f"=IF({N}{excel_row}=\"连续焊\",{O}{excel_row}*{cref('连')},"
            f"IF({N}{excel_row}=\"点焊\",{O}{excel_row}*{cref('点')},0))"
        ))
        ws.cell(row=excel_row, column=col_index["焊接标准件费(元)"], value=f"={R}{excel_row}*{cref('标')}*{I}{excel_row}")
        ws.cell(row=excel_row, column=col_index["攻丝费(元)"], value=(
            f"=IF(OR(ISNUMBER(SEARCH(\"5052\",{E}{excel_row})),ISNUMBER(SEARCH(\"铝\",{E}{excel_row}))),"
            f"{U}{excel_row}*{cref('丝铝')}*{I}{excel_row},{U}{excel_row}*{cref('丝他')}*{I}{excel_row})"
        ))
        ws.cell(row=excel_row, column=col_index["表面处理费(元)"], value=(
            f"=IF({W}{excel_row}=\"喷粉\",{AD}{excel_row}*{AE}{excel_row}*2/1000000*{cref('喷')},"
            f"IF({W}{excel_row}=\"磷化\",{AD}{excel_row}*{AE}{excel_row}*2/1000000*{cref('磷')},0))"
        ))
        ws.cell(row=excel_row, column=col_index["单重(kg)"], value=(
            f"={AD}{excel_row}*{AE}{excel_row}*{H}{excel_row}*"
            f"IF({E}{excel_row}=\"5052铝板\",2.73,IF(OR({E}{excel_row}=\"Q235冷轧钢板\",{E}{excel_row}=\"镀锌板\"),7.85,"
            f"IF({E}{excel_row}=\"SUS304不锈钢\",7.93,7.85)))/1000000"
        ))
        ws.cell(row=excel_row, column=col_index["总重(kg)"], value=f"={Y}{excel_row}*{I}{excel_row}")
        ws.cell(row=excel_row, column=col_index["材料费(元)"], value=(
            f"={Z}{excel_row}*IF({E}{excel_row}=\"5052铝板\",24,IF(OR({E}{excel_row}=\"Q235冷轧钢板\",{E}{excel_row}=\"镀锌板\"),5,"
            f"IF({E}{excel_row}=\"SUS304不锈钢\",18,5)))*(1+{cref('损耗')})"
        ))
        ws.cell(row=excel_row, column=col_index["单品含税总价(元)"], value=(
            f"=({AA}{excel_row}+{K}{excel_row}+{M}{excel_row}+{P}{excel_row}+{S}{excel_row}+{V}{excel_row}+{X}{excel_row})*(1+{cref('税率')})"
        ))
        ws.cell(row=excel_row, column=col_index["焊接输入提示"], value=(
            f"=IF({N}{excel_row}=\"连续焊\",\"请输入焊接长度(mm)\","
            f"IF({N}{excel_row}=\"点焊\",\"请输入焊点数量\",\"无焊接，无需填写\"))"
        ))

    if n > 0:
        sum_row = data_start + n
        ab_letter = L("单品含税总价(元)")
        ws.cell(row=sum_row, column=1, value="汇总")
        ws.cell(
            row=sum_row,
            column=col_index["单品含税总价(元)"],
            value=f"=SUM({ab_letter}{data_start}:{ab_letter}{data_start + n - 1})",
        )
        for c in (1, col_index["单品含税总价(元)"]):
            cell = ws.cell(row=sum_row, column=c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill

    for idx, h in enumerate(headers, start=1):
        if h in ("焊接标准件名称", "攻丝规格"):
            w = 12
        else:
            w = 11
        ws.column_dimensions[get_column_letter(idx)].width = w

    # ---------- Sheet：使用说明 ----------
    ws_h = wb.create_sheet("使用说明")
    ws_h["A1"] = (
        "1）「报价明细」中灰色表头列为自动公式，修改长宽厚数量等输入列后会自动重算。\n"
        "2）所有可调单价、税率、损耗集中在「单价参数」表；切割/折弯/焊接/攻丝/表面处理系数均可在此改。\n"
        "3）材质对应的密度与材料单价如需扩展，可在「材质密度与单价」表增行，并同步调整「单重」「材料费」列公式中的 IF 分支。\n"
        "4）本文件与网页新星模板规则一致；网页侧修改了 MATERIAL_PARAMS 时，请对照更新本表或重新导出。"
    )
    ws_h["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_h.column_dimensions["A"].width = 100

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ----- 崂应：参数中文名 -----
_LY_PARAM_LABELS: Dict[str, str] = {
    "cutting_rate": "切割费率（米数×厚×该系数）",
    "grinding_small": "打磨基数（长或宽<200mm，每件）",
    "grinding_large": "打磨基数（否则，每件）",
    "grinding_mult": "打磨倍数（×加工数量）",
    "pierce_per_t": "穿孔系数（×厚×穿孔数量）",
    "weld_a": "焊接长度系数A",
    "weld_b": "焊接长度系数B（焊接费相关）",
    "countersink": "沉孔单价（元/个）",
    "bend_al": "折弯单价-铝板（小件，元/次·件）",
    "bend_steel": "折弯单价-钢/不锈钢（小件）",
    "bend_large": "折弯单价-大件（长或宽>100）",
    "pem_a": "压铆系数A",
    "pem_b": "压铆系数B",
    "tap_al": "攻丝单价-铝板（元/个）",
    "tap_steel": "攻丝单价-钢/不锈钢（元/个）",
}


_LY_FORMULA_NOTES: List[Tuple[str, str]] = [
    ("重量(kg)", "单件毛重：由长×宽×厚×材质密度÷1e6×总重系数得到（总重系数见下方快照旁参数）。"),
    ("总重(kg)", "重量(kg)×加工数量。"),
    ("打磨(元)", "按长宽是否小于200mm取打磨基数×加工数量×打磨倍数。"),
    ("材料费(元)", "与总重、材料单价及脚本内材料计费规则一致（详见网页计算）。"),
    ("总加工费(元)", "含切割、穿孔、焊接、沉孔、折弯、压铆、攻丝、激光打标等脚本汇总项。"),
    ("表面处理费(元)", "按所选表面处理类型与面积/单价规则计算。"),
    ("含税单价(元/件)", "单件含税价格（含材料+加工+表处等后含税摊算）。"),
    ("含税总价(元)", "含税单价×加工数量。"),
    ("总重系数", "侧边栏「总重系数」：理重×系数计入毛重，用于后续按重量计费列。"),
]


def build_laoying_quote_detail_workbook_bytes(
    df: pd.DataFrame,
    price_params: Mapping[str, float],
    weight_coef: float,
    material_library: Mapping[str, Mapping[str, Any]],
) -> bytes:
    """
    崂应：明细数值快照 + 当前加工单价 + 材质库 + 公式文字说明。
    """
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "明细快照"
    ws0.append([f"导出时间 {datetime.now():%Y-%m-%d %H:%M:%S}"])
    ws0.append([f"总重系数（与侧边栏一致）", float(weight_coef)])
    ws0.append(list(df.columns))
    for _, row in df.iterrows():
        ws0.append([row.get(c, "") for c in df.columns])
    ws0.freeze_panes = "A4"
    for cell in ws0[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1E3A8A")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    ws1 = wb.create_sheet("加工单价参数")
    ws1.append(["参数键", "当前数值", "含义（中文）"])
    for k, v in price_params.items():
        ws1.append([k, float(v), _LY_PARAM_LABELS.get(k, "")])
    ws1.append(["weight_coef_snapshot", float(weight_coef), "导出时的总重系数快照（与明细快照表头旁一致）"])
    for cell in ws1[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1E3A8A")

    ws2 = wb.create_sheet("材料库")
    ws2.append(["材质", "密度", "材料单价(元/kg)", "类别"])
    for name, info in material_library.items():
        ws2.append(
            [
                name,
                info.get("density", ""),
                info.get("unit_price", ""),
                info.get("category", ""),
            ]
        )
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1E3A8A")

    ws3 = wb.create_sheet("计算列公式说明")
    ws3.append(["项目", "说明（与网页脚本一致，供线下改表对照）"])
    for name, note in _LY_FORMULA_NOTES:
        ws3.append([name, note])
    for cell in ws3[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1E3A8A")
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 80

    for sheet in (ws0, ws1, ws2):
        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    v = cell.value
                    if v is not None and len(str(v)) > max_length:
                        max_length = len(str(v))
                except Exception:
                    pass
            sheet.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 45)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_placeholder_quote_workbook_bytes(
    *,
    customer_name: str,
    columns: Sequence[str],
    note: str,
) -> bytes:
    """常规/鼎信占位：导出列头模板 + 说明。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "模板说明"
    ws["A1"] = f"【{customer_name}】{note}"
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 90

    wt = wb.create_sheet("报价明细模板")
    wt.append(list(columns))
    for cell in wt[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1E3A8A")
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
